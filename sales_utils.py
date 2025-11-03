# sales_utils.py (upgraded for cart-based transactions and voiding)
import sqlite3
from datetime import datetime, timedelta
from fpdf import FPDF
import hashlib
import os
import json
import bcrypt
from typing import List, Dict, Tuple, Optional

DB_NAME = "bar_sales.db"

# =========================
# Database initialization and migration
# =========================

def _table_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    cur = conn.cursor()
    try:
        cur.execute(f"PRAGMA table_info({table})")
        return [row[1] for row in cur.fetchall()]
    except sqlite3.OperationalError:
        return []

def ensure_cart_schema():
    """Ensure new cart-based schema exists. If legacy 'sales' exists with item columns,
    rename to 'sales_legacy' and create new 'sales' header + 'sale_items' tables."""
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        # Detect legacy 'sales' table (row-level schema)
        legacy_sales_cols = _table_columns(conn, 'sales')
        legacy_signature = {'username', 'item', 'quantity', 'price_per_unit', 'total', 'timestamp'}
        is_legacy = legacy_signature.issubset(set(legacy_sales_cols))

        # Rename legacy table if not yet migrated
        if is_legacy:
            cur.execute("ALTER TABLE sales RENAME TO sales_legacy")

        # Create new header table 'sales'
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                transaction_id TEXT UNIQUE NOT NULL,
                cashier TEXT NOT NULL,
                total REAL NOT NULL,
                timestamp TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'ACTIVE',
                void_reason TEXT DEFAULT NULL,
                void_authorized_by TEXT DEFAULT NULL,
                voided_at TEXT DEFAULT NULL
            )
            """
        )

        # Create items table
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS sale_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER NOT NULL,
                item TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                unit_price REAL NOT NULL,
                subtotal REAL NOT NULL,
                FOREIGN KEY (sale_id) REFERENCES sales(id) ON DELETE CASCADE
            )
            """
        )

        # Indexes
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sales_ts ON sales(timestamp)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sale_items_sale_id ON sale_items(sale_id)")

        conn.commit()
    finally:
        conn.close()

def init_db():
    """Initialize DB (legacy tables for backward compat) and ensure cart-based schema."""
    conn = sqlite3.connect(DB_NAME, timeout=30)
    cur = conn.cursor()
    
    # Enable WAL mode for better concurrency
    cur.execute("PRAGMA journal_mode=WAL")
    cur.execute("PRAGMA synchronous=NORMAL")
    cur.execute("PRAGMA cache_size=10000")
    cur.execute("PRAGMA temp_store=memory")
    
    # Legacy row-level sales table (kept for compat; may be renamed by ensure_cart_schema)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            item TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            price_per_unit REAL NOT NULL,
            total REAL NOT NULL,
            timestamp TEXT NOT NULL
        )
    ''')
    # Users
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL
        )
    ''')
    # Inventory
    cur.execute('''
        CREATE TABLE IF NOT EXISTS inventory (
            item TEXT PRIMARY KEY,
            quantity INTEGER NOT NULL,
            cost_price REAL DEFAULT 0,
            selling_price REAL DEFAULT 0,
            category TEXT DEFAULT ''
        )
    ''')
    # Notes table for communication
    cur.execute('''
        CREATE TABLE IF NOT EXISTS notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sender TEXT NOT NULL,
            receiver TEXT NOT NULL,
            message TEXT NOT NULL,
            timestamp TEXT NOT NULL,
            is_read INTEGER DEFAULT 0
        )
    ''')
    conn.commit()
    conn.close()

    # Create/migrate to cart schema
    ensure_cart_schema()

# =========================
# Auth helpers
# =========================

def hash_password(password):
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def check_password(password, password_hash):
    return bcrypt.checkpw(password.encode(), password_hash.encode())

def create_user(username, password, role):
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    password_hash = hash_password(password)
    try:
        cur.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)", (username, password_hash, role))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return False
    finally:
        conn.close()
    return True

def get_user(username):
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute("SELECT username, password_hash, role FROM users WHERE username=?", (username,))
        row = cur.fetchone()
    finally:
        conn.close()
    if row:
        return {'username': row[0], 'password_hash': row[1], 'role': row[2]}
    return None

# =========================
# Inventory helpers
# =========================

def get_stock(item):
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('SELECT quantity FROM inventory WHERE item=?', (item,))
        row = cur.fetchone()
    finally:
        conn.close()
    return row[0] if row else 0

def update_stock(item, change):
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('SELECT quantity FROM inventory WHERE item=?', (item,))
        row = cur.fetchone()
        if row:
            new_qty = row[0] + change
            cur.execute('UPDATE inventory SET quantity=? WHERE item=?', (new_qty, item))
        else:
            cur.execute('INSERT INTO inventory (item, quantity) VALUES (?, ?)', (item, max(0, change)))
        conn.commit()
    finally:
        conn.close()

def get_all_stock():
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('SELECT item, quantity, category FROM inventory ORDER BY item')
        rows = cur.fetchall()
    finally:
        conn.close()
    return rows

def set_item_prices(item, cost, sell):
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        if cost is not None and sell is not None:
            cur.execute('UPDATE inventory SET cost_price=?, selling_price=? WHERE item=?', (cost, sell, item))
        elif cost is not None:
            cur.execute('UPDATE inventory SET cost_price=? WHERE item=?', (cost, item))
        elif sell is not None:
            cur.execute('UPDATE inventory SET selling_price=? WHERE item=?', (sell, item))
        conn.commit()
    finally:
        conn.close()

def set_item_category(item, category):
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('UPDATE inventory SET category=? WHERE item=?', (category, item))
        conn.commit()
    finally:
        conn.close()

def get_item_category(item):
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('SELECT category FROM inventory WHERE item=?', (item,))
        row = cur.fetchone()
    finally:
        conn.close()
    return row[0] if row else ''

def get_categories():
    """Get all unique categories from inventory"""
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('SELECT DISTINCT category FROM inventory WHERE category IS NOT NULL AND category != "" ORDER BY category')
        rows = cur.fetchall()
        return [row[0] for row in rows]
    finally:
        conn.close()

def delete_item(item):
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('DELETE FROM inventory WHERE item=?', (item,))
        conn.commit()
    finally:
        conn.close()

def get_item_prices(item) -> Tuple[Optional[float], Optional[float]]:
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('SELECT cost_price, selling_price FROM inventory WHERE item=?', (item,))
        row = cur.fetchone()
    finally:
        conn.close()
    return row if row else (None, None)

# =========================
# Notes/Communication System
# =========================

def send_note(sender, receiver, message):
    """Send a note from sender to receiver"""
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        cur.execute("INSERT INTO notes (sender, receiver, message, timestamp) VALUES (?, ?, ?, ?)",
                    (sender, receiver, message, timestamp))
        conn.commit()
        return True
    except Exception as e:
        print(f"Error sending note: {e}")
        return False
    finally:
        conn.close()

def get_notes_for_user(username, unread_only=False):
    """Get notes for a user"""
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        if unread_only:
            cur.execute("SELECT id, sender, message, timestamp FROM notes WHERE receiver=? AND is_read=0 ORDER BY timestamp DESC",
                        (username,))
        else:
            cur.execute("SELECT id, sender, message, timestamp, is_read FROM notes WHERE receiver=? ORDER BY timestamp DESC",
                        (username,))
        rows = cur.fetchall()
        return rows
    finally:
        conn.close()

def mark_note_as_read(note_id):
    """Mark a note as read"""
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute("UPDATE notes SET is_read=1 WHERE id=?", (note_id,))
        conn.commit()
        return True
    except Exception as e:
        print(f"Error marking note as read: {e}")
        return False
    finally:
        conn.close()

# =========================
# Legacy single-line sale API (compat mode)
# =========================

def record_sale(username, item, quantity, price_per_unit):
    """Compatibility layer. If cart tables exist, record as a one-item sale. If not, write to legacy table."""
    # Stock check
    stock = get_stock(item)
    if quantity > stock:
        raise ValueError(f"Not enough stock for {item}. In stock: {stock}, requested: {quantity}")

    conn = sqlite3.connect(DB_NAME, timeout=10)
    has_header = 'transaction_id' in _table_columns(conn, 'sales')
    has_items = len(_table_columns(conn, 'sale_items')) > 0
    conn.close()

    if has_header and has_items:
        # Create cart sale with one item
        _, _, total = create_sale_with_items(username, [{
            'item': item,
            'quantity': quantity,
            'unit_price': price_per_unit
        }])
        return total
    else:
        # Legacy write
        total = quantity * price_per_unit
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn2 = sqlite3.connect(DB_NAME, timeout=10)
        cur2 = conn2.cursor()
        cur2.execute("INSERT INTO sales (username, item, quantity, price_per_unit, total, timestamp) VALUES (?, ?, ?, ?, ?, ?)",
                     (username, item, quantity, price_per_unit, total, timestamp))
        conn2.commit()
        conn2.close()
        update_stock(item, -quantity)
        return total

# =========================
# Cart-based sale API
# =========================

def _new_transaction_id() -> str:
    dt = datetime.now().strftime('%Y%m%d%H%M%S')
    rnd = hashlib.sha256(os.urandom(16)).hexdigest()[:6]
    return f"TX-{dt}-{rnd}"

def create_sale_with_items(cashier: str, items: List[Dict]) -> Tuple[int, str, float]:
    """Create sale header + multiple sale_items with proper transaction handling.
    items: list of dicts with keys: item, quantity, unit_price
    Returns (sale_id, transaction_id, total)
    """
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=30)
    cur = conn.cursor()
    
    try:
        # Begin immediate transaction to prevent locks
        cur.execute("BEGIN IMMEDIATE")
        
        # Stock check within transaction
        for it in items:
            cur.execute('SELECT quantity FROM inventory WHERE item=?', (it['item'],))
            row = cur.fetchone()
            stock = row[0] if row else 0
            if it['quantity'] > stock:
                raise ValueError(f"Not enough stock for {it['item']}. In stock: {stock}, requested: {it['quantity']}")
        
        # Compute total
        total = sum(float(it['quantity']) * float(it['unit_price']) for it in items)
        tx_id = _new_transaction_id()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Insert sale header
        cur.execute("INSERT INTO sales (transaction_id, cashier, total, timestamp, status) VALUES (?, ?, ?, ?, 'ACTIVE')",
                    (tx_id, cashier, total, ts))
        sale_id = cur.lastrowid
        
        # Insert sale items and update stock in same transaction
        for it in items:
            subtotal = float(it['quantity']) * float(it['unit_price'])
            
            # Insert sale item
            cur.execute(
                "INSERT INTO sale_items (sale_id, item, quantity, unit_price, subtotal) VALUES (?, ?, ?, ?, ?)",
                (sale_id, it['item'], int(it['quantity']), float(it['unit_price']), subtotal)
            )
            
            # Update stock within same transaction
            cur.execute('SELECT quantity FROM inventory WHERE item=?', (it['item'],))
            row = cur.fetchone()
            if row:
                new_qty = row[0] - int(it['quantity'])
                cur.execute('UPDATE inventory SET quantity=? WHERE item=?', (new_qty, it['item']))
        
        # Commit transaction
        conn.commit()
        return sale_id, tx_id, total
        
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

def void_sale_transaction(sale_id: int, reason: str, requested_by: str, authorized_by: str) -> Tuple[bool, str]:
    """Void entire sale, restore stock, mark as VOIDED with audit fields."""
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=30)
    cur = conn.cursor()
    
    try:
        cur.execute("BEGIN IMMEDIATE")
        
        cur.execute("SELECT status FROM sales WHERE id=?", (sale_id,))
        row = cur.fetchone()
        if not row:
            return False, "Sale not found"
        if row[0] == 'VOIDED':
            return False, "Sale already voided"
        
        # Restore stock for all items within same transaction
        cur.execute("SELECT item, quantity FROM sale_items WHERE sale_id=?", (sale_id,))
        for item, qty in cur.fetchall():
            cur.execute('SELECT quantity FROM inventory WHERE item=?', (item,))
            row = cur.fetchone()
            if row:
                new_qty = row[0] + int(qty)
                cur.execute('UPDATE inventory SET quantity=? WHERE item=?', (new_qty, item))
        
        # Mark as voided
        cur.execute(
            "UPDATE sales SET status='VOIDED', void_reason=?, void_authorized_by=?, voided_at=? WHERE id=?",
            (reason, authorized_by, datetime.now().isoformat(timespec='seconds'), sale_id)
        )
        
        conn.commit()
        return True, "Sale voided successfully"
        
    except Exception as e:
        conn.rollback()
        return False, f"Error voiding sale: {e}"
    finally:
        conn.close()

# =========================
# Queries and reporting
# =========================

def get_recent_sales_headers(limit: int = 20) -> List[Tuple]:
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, transaction_id, cashier, total, timestamp, status FROM sales ORDER BY timestamp DESC LIMIT ?",
            (limit,)
        )
        rows = cur.fetchall()
    finally:
        conn.close()
    return rows

def get_sale_items(sale_id: int) -> List[Tuple]:
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT item, quantity, unit_price, subtotal FROM sale_items WHERE sale_id=? ORDER BY id",
            (sale_id,)
        )
        rows = cur.fetchall()
    finally:
        conn.close()
    return rows

def get_daily_summary(day: Optional[str] = None) -> Dict:
    ensure_cart_schema()
    if day is None:
        day = datetime.now().strftime('%Y-%m-%d')
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute("SELECT COALESCE(SUM(total),0) FROM sales WHERE status!='VOIDED' AND DATE(timestamp)=?", (day,))
        total = cur.fetchone()[0] or 0
        cur.execute(
            """
            SELECT si.item, SUM(si.quantity) as qty
            FROM sale_items si
            JOIN sales s ON s.id=si.sale_id
            WHERE s.status!='VOIDED' AND DATE(s.timestamp)=?
            GROUP BY si.item
            ORDER BY qty DESC
            LIMIT 5
            """,
            (day,)
        )
        top_items = cur.fetchall()
        cur.execute(
            """
            SELECT cashier, COALESCE(SUM(total),0) as tot
            FROM sales
            WHERE status!='VOIDED' AND DATE(timestamp)=?
            GROUP BY cashier
            ORDER BY tot DESC
            """,
            (day,)
        )
        cashier_perf = cur.fetchall()
    finally:
        conn.close()
    return {
        'date': day,
        'total_sales': total,
        'top_items': top_items,
        'cashier_performance': cashier_perf
    }

def get_weekly_summary() -> Dict:
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        today = datetime.now().date()
        start = (today - timedelta(days=6)).strftime('%Y-%m-%d')
        end = today.strftime('%Y-%m-%d')
        cur.execute(
            """
            SELECT DATE(timestamp) d, COALESCE(SUM(total),0)
            FROM sales
            WHERE status!='VOIDED' AND DATE(timestamp) BETWEEN ? AND ?
            GROUP BY DATE(timestamp)
            ORDER BY DATE(timestamp)
            """,
            (start, end)
        )
        rows = cur.fetchall()
    finally:
        conn.close()
    return {'range': (start, end), 'daily_totals': rows}

# =========================
# Exports and backup
# =========================

def export_to_csv():
    import csv
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM sales")
        rows = cur.fetchall()
        headers = [d[0] for d in cur.description]
    finally:
        conn.close()

    if not os.path.exists("exports"):
        os.makedirs("exports")

    # Keep filename as sales.csv for compatibility
    with open("exports/sales.csv", "w", newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(headers)
        writer.writerows(rows)

def get_total_sales():
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute("SELECT COALESCE(SUM(total),0) FROM sales WHERE status!='VOIDED'")
        total = cur.fetchone()[0]
    finally:
        conn.close()
    return total if total else 0

# =========================
# Receipts
# =========================

def generate_pdf_receipt(receipt_id, username, item, quantity, price_per_unit, total, timestamp):
    """Legacy single-item receipt (kept for compatibility with existing UI)."""
    business_name = "Comfort_2022 Bar Sales"
    business_address = "123 Main St, Your City"
    logo_path = os.path.join("assets", "logo.png")
    signature_data = f"{receipt_id}|{username}|{item}|{quantity}|{price_per_unit}|{total}|{timestamp}"
    signature = hashlib.sha256(signature_data.encode()).hexdigest()
    pdf = FPDF()
    pdf.add_page()
    if os.path.exists(logo_path):
        try:
            pdf.image(logo_path, x=10, y=8, w=33)
            pdf.set_xy(50, 10)
        except Exception:
            pdf.set_xy(10, 10)
    else:
        pdf.set_xy(10, 10)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, business_name, ln=1)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, business_address, ln=1)
    pdf.ln(10)
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, f"Receipt ID: {receipt_id}", ln=1)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Date/Time: {timestamp}", ln=1)
    pdf.cell(0, 10, f"Cashier: {username}", ln=1)
    pdf.ln(5)
    pdf.cell(0, 10, f"Item: {item}", ln=1)
    pdf.cell(0, 10, f"Quantity: {quantity}", ln=1)
    pdf.cell(0, 10, f"Price per Unit: ZMW {price_per_unit:.2f}", ln=1)
    pdf.cell(0, 10, f"Total: ZMW {total:.2f}", ln=1)
    pdf.ln(10)
    pdf.set_font("Arial", "I", 10)
    pdf.multi_cell(0, 8, f"Digital Signature:\n{signature}")
    if not os.path.exists("exports"):
        os.makedirs("exports")
    pdf_path = os.path.join("exports", f"receipt_{receipt_id}.pdf")
    pdf.output(pdf_path)
    return pdf_path

def generate_pdf_receipt_for_sale(sale_id: int) -> str:
    """Generate a multi-item receipt for a sale header+items."""
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute("SELECT transaction_id, cashier, total, timestamp, status FROM sales WHERE id=?", (sale_id,))
        row = cur.fetchone()
        if not row:
            raise ValueError("Sale not found")
        tx_id, cashier, total, ts, status = row
        cur.execute("SELECT item, quantity, unit_price, subtotal FROM sale_items WHERE sale_id=? ORDER BY id", (sale_id,))
        items = cur.fetchall()
    finally:
        conn.close()

    business_name = "Comfort_2022 Bar Sales"
    business_address = "123 Main St, Your City"
    logo_path = os.path.join("assets", "logo.png")
    signature_data = f"{tx_id}|{cashier}|{total}|{ts}|{len(items)}"
    signature = hashlib.sha256(signature_data.encode()).hexdigest()

    pdf = FPDF()
    pdf.add_page()
    if os.path.exists(logo_path):
        try:
            pdf.image(logo_path, x=10, y=8, w=33)
            pdf.set_xy(50, 10)
        except Exception:
            pdf.set_xy(10, 10)
    else:
        pdf.set_xy(10, 10)

    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, business_name, ln=1)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, business_address, ln=1)
    pdf.ln(5)

    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, f"Receipt: {tx_id}", ln=1)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 8, f"Date/Time: {ts}", ln=1)
    pdf.cell(0, 8, f"Cashier: {cashier}", ln=1)
    pdf.cell(0, 8, f"Status: {status}", ln=1)
    pdf.ln(5)

    # Table header
    pdf.set_font("Arial", "B", 12)
    pdf.cell(80, 8, "Item", 1)
    pdf.cell(25, 8, "Qty", 1, 0, 'R')
    pdf.cell(35, 8, "Unit", 1, 0, 'R')
    pdf.cell(40, 8, "Subtotal", 1, 1, 'R')
    pdf.set_font("Arial", "", 12)

    for it, qty, unit, sub in items:
        pdf.cell(80, 8, str(it), 1)
        pdf.cell(25, 8, str(qty), 1, 0, 'R')
        pdf.cell(35, 8, f"{unit:.2f}", 1, 0, 'R')
        pdf.cell(40, 8, f"{sub:.2f}", 1, 1, 'R')

    pdf.ln(3)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(140, 8, "TOTAL", 1)
    pdf.cell(40, 8, f"{total:.2f}", 1, 1, 'R')

    pdf.ln(8)
    pdf.set_font("Arial", "I", 9)
    pdf.multi_cell(0, 6, f"Digital Signature:\n{signature}")

    if not os.path.exists("exports"):
        os.makedirs("exports")
    pdf_path = os.path.join("exports", f"receipt_{tx_id}.pdf")
    pdf.output(pdf_path)
    return pdf_path

# =========================
# Backups, audit, exports
# =========================

def backup_today_sales():
    ensure_cart_schema()
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM sales WHERE DATE(timestamp)=?", (today,))
        rows = cur.fetchall()
        columns = [desc[0] for desc in cur.description]
    finally:
        conn.close()
    sales_list = [dict(zip(columns, row)) for row in rows]
    if not os.path.exists("data"):
        os.makedirs("data")
    backup_path = os.path.join("data", f"backup_{today}.json")
    with open(backup_path, "w", encoding="utf-8") as f:
        json.dump(sales_list, f, indent=2)
    return backup_path

def log_audit_event(event):
    if not os.path.exists("data"):
        os.makedirs("data")
    log_path = os.path.join("data", "audit.log")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {event}\n")

def export_all_sales_to_csv(start_date, end_date, selected_columns, export_format='CSV'):
    import csv
    ensure_cart_schema()
    # Validate against new header columns
    valid_cols = { 'id','transaction_id','cashier','total','timestamp','status','void_reason','void_authorized_by','voided_at' }
    cols = [c for c in selected_columns if c in valid_cols]
    if not cols:
        cols = ['id','transaction_id','cashier','total','timestamp','status']
    col_str = ', '.join(cols)
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        query = f"SELECT {col_str} FROM sales WHERE DATE(timestamp) BETWEEN ? AND ?"
        cur.execute(query, (start_date, end_date))
        rows = cur.fetchall()
    finally:
        conn.close()

    if not os.path.exists("exports"):
        os.makedirs("exports")
    today = datetime.now().strftime('%Y-%m-%d')

    if export_format == 'Excel':
        import openpyxl
        from openpyxl.utils import get_column_letter
        xlsx_path = os.path.join("exports", f"all_sales_{start_date}_to_{end_date}_{today}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(cols)
        for row in rows:
            ws.append(row)
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)
        wb.save(xlsx_path)
        return xlsx_path
    else:
        csv_path = os.path.join("exports", f"all_sales_{start_date}_to_{end_date}_{today}.csv")
        with open(csv_path, "w", newline='', encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(cols)
            writer.writerows(rows)
        return csv_path

# =========================
# Item sales history (cart tables)
# =========================

def get_sales_history_for_item(item):
    ensure_cart_schema()
    conn = sqlite3.connect(DB_NAME, timeout=10)
    cur = conn.cursor()
    try:
        cur.execute('''
            SELECT s.timestamp, s.cashier, si.quantity, si.unit_price, si.subtotal
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE si.item = ?
            ORDER BY s.timestamp DESC
        ''', (item,))
        rows = cur.fetchall()
    finally:
        conn.close()
    return rows
